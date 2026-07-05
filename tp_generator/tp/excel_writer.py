"""手間いらず読み込み用Excelファイルの出力（仕様書 2〜3章）。

【最重要】手間いらずソフトウェアは、配列構造が1セルでもずれると
読み込みに失敗する。以下の配列を完全に遵守する:

  行N   : A=部屋タイプ        B=[部屋タイプ名]
  行N+1 : A=除外部屋タイプ    B=[除外部屋タイプ名]
  行N+2 : A=ベース料金ランク  B=[料金ランク]
  行N+3 : （空行）
  行N+4 :                     B=→残室数  C=※手仕舞いの場合...
  行N+5 : A=↓何日前の状況か   B=日数/残室数  C以降=[残室軸]
  行N+6〜:                    B=[リードタイム]  C以降=[料金ランク]

各マトリクス間は必ず2行空ける。
"""

from __future__ import annotations

from openpyxl import Workbook

from .engine import Matrix
from .master import Master

INTRO_TEXT = [
    "このファイルは「手間いらず」ターゲットプライス（TP）設定用のDATAシートを含みます。",
    "",
    "・DATAシートが主要編集対象です。MASTERシートは料金ランク一覧の定義です。",
    "・配列構造が1セルでもずれると手間いらずは読み込みに失敗します。",
    "・各マトリクス間は必ず2行空けてください。",
    "・余計な空白セル、改行を入れないでください（全角/半角スペースに注意）。",
]


def write_workbook(matrices: list[Matrix], master: Master, path: str) -> None:
    """はじめに / MASTER / SETTING / DATA の4シート構成でxlsxを出力する。"""
    wb = Workbook()

    ws_intro = wb.active
    ws_intro.title = "はじめに"
    for i, line in enumerate(INTRO_TEXT, start=1):
        ws_intro.cell(row=i, column=1, value=line)

    ws_master = wb.create_sheet("MASTER")
    ws_master.cell(row=1, column=1, value="料金ランク")
    for i, rank in enumerate(master.ranks, start=2):
        ws_master.cell(row=i, column=1, value=str(rank))

    ws_setting = wb.create_sheet("SETTING")
    ws_setting.cell(row=1, column=1, value="設定項目")
    ws_setting.cell(row=1, column=2, value="値")

    ws_data = wb.create_sheet("DATA")
    row = 1
    for m in matrices:
        row = _write_matrix(ws_data, row, m)
        row += 2  # マトリクス間は必ず2行空ける

    wb.save(path)


def _write_matrix(ws, start_row: int, m: Matrix) -> int:
    """1マトリクスを書き込み、最終行の次の行番号を返す。"""
    r = start_row
    ws.cell(row=r, column=1, value="部屋タイプ")
    ws.cell(row=r, column=2, value=m.room_type)
    r += 1
    ws.cell(row=r, column=1, value="除外部屋タイプ")
    if m.excluded_room_type:
        ws.cell(row=r, column=2, value=m.excluded_room_type)
    r += 1
    ws.cell(row=r, column=1, value="ベース料金ランク")
    ws.cell(row=r, column=2, value=str(m.base_rank))
    r += 1
    # 空行
    r += 1
    ws.cell(row=r, column=2, value="→残室数")
    if m.stock_note:
        ws.cell(row=r, column=3, value=m.stock_note)
    r += 1
    ws.cell(row=r, column=1, value="↓何日前の状況か")
    ws.cell(row=r, column=2, value="日数/残室数")
    for c, stock in enumerate(m.stock_axis, start=3):
        ws.cell(row=r, column=c, value=stock)
    r += 1
    for lead, cells in zip(m.lead_times, m.cells):
        ws.cell(row=r, column=2, value=lead)
        for c, rank in enumerate(cells, start=3):
            ws.cell(row=r, column=c, value=str(rank))
        r += 1
    return r
