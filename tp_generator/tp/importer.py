"""既存の手間いらずTP設定ファイル（元ファイル）のDATAシートを解析する。

作業フロー Step 1（元ファイル分析）に対応:
  - 既存マトリクスの構造を把握
  - ルールの種類を特定（残室軸・リードタイム軸の組み合わせで分類）
  - 各ルールのベース料金ランク・軸・全データ値を記録
"""

from __future__ import annotations

from openpyxl import load_workbook

from .engine import Matrix
from .master import parse_rank


def read_data_sheet(path: str, sheet_name: str = "DATA") -> list[Matrix]:
    """DATAシートから全マトリクスを読み取る。"""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート {sheet_name!r} が見つかりません（存在: {wb.sheetnames}）")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    matrices: list[Matrix] = []
    i = 0
    while i < len(rows):
        a = _cell(rows, i, 0)
        if a == "部屋タイプ":
            m, i = _read_matrix(rows, i)
            matrices.append(m)
        else:
            i += 1
    if not matrices:
        raise ValueError("DATAシートにマトリクスが見つかりませんでした")
    return matrices


def _cell(rows, r: int, c: int):
    if r >= len(rows) or c >= len(rows[r]):
        return None
    v = rows[r][c]
    return v.strip() if isinstance(v, str) else v


def _read_matrix(rows, start: int) -> tuple[Matrix, int]:
    r = start
    room_type = str(_cell(rows, r, 1) or "")
    r += 1
    if _cell(rows, r, 0) != "除外部屋タイプ":
        raise ValueError(f"行{r + 1}: 「除外部屋タイプ」がありません（配列構造ずれ）")
    excluded = str(_cell(rows, r, 1) or "")
    r += 1
    if _cell(rows, r, 0) != "ベース料金ランク":
        raise ValueError(f"行{r + 1}: 「ベース料金ランク」がありません（配列構造ずれ）")
    base_rank = parse_rank(str(_cell(rows, r, 1)))
    r += 1

    # ヘッダー行（↓何日前の状況か）を探す（空行・→残室数行を読み飛ばす）
    stock_note = ""
    while r < len(rows) and _cell(rows, r, 0) != "↓何日前の状況か":
        if _cell(rows, r, 1) == "→残室数":
            stock_note = str(_cell(rows, r, 2) or "")
        r += 1
    if r >= len(rows):
        raise ValueError(f"行{start + 1}からのマトリクスに「↓何日前の状況か」行がありません")

    stock_axis: list[int] = []
    c = 2
    while _cell(rows, r, c) is not None and str(_cell(rows, r, c)) != "":
        stock_axis.append(int(_cell(rows, r, c)))
        c += 1
    r += 1

    lead_times: list[int] = []
    cells: list[list] = []
    while r < len(rows):
        lead = _cell(rows, r, 1)
        if lead is None or str(lead) == "":
            break
        lead_times.append(int(lead))
        row_cells = [parse_rank(str(_cell(rows, r, 2 + ci))) for ci in range(len(stock_axis))]
        cells.append(row_cells)
        r += 1

    m = Matrix(
        room_type=room_type,
        base_rank=base_rank,
        lead_times=lead_times,
        stock_axis=stock_axis,
        cells=cells,
        excluded_room_type=excluded,
        stock_note=stock_note,
    )
    return m, r


def to_config(matrices: list[Matrix], master_strings: list[str] | None = None) -> dict:
    """マトリクス一覧からルールタイプを特定し、生成用の設定dictを作る。

    残室軸・リードタイム軸が同じマトリクスを1つのルールタイプとみなし、
    そのグループの最初のマトリクスをベースとして採用する。
    適用範囲はグループ内のベース料金ランクの最小〜最大とする。
    """
    groups: dict[tuple, list[Matrix]] = {}
    for m in matrices:
        key = (tuple(m.lead_times), tuple(m.stock_axis))
        groups.setdefault(key, []).append(m)

    rules = []
    for idx, ms in enumerate(groups.values(), start=1):
        base = ms[0]
        ranks = sorted(m.base_rank for m in ms)
        rules.append(
            {
                "name": f"ルール{idx}（{len(base.lead_times)}段階×{len(base.stock_axis)}段階）",
                "base_rank": str(base.base_rank),
                "apply_from": str(ranks[0]),
                "apply_to": str(ranks[-1]),
                "lead_times": list(base.lead_times),
                "stock_axis": list(base.stock_axis),
                "matrix": [[str(c) for c in row] for row in base.cells],
            }
        )

    config: dict = {
        "room_type": matrices[0].room_type,
        "excluded_room_type": matrices[0].excluded_room_type,
        "rules": rules,
    }
    if master_strings:
        config["master"] = master_strings
    return config


def read_master_sheet(path: str, sheet_name: str = "MASTER") -> list[str] | None:
    """MASTERシートから料金ランク一覧を読み取る（無ければNone）。"""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    ranks: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str) and v.strip().startswith("("):
                ranks.append(v.strip())
    return ranks or None
