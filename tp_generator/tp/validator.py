"""生成ファイルの検証（仕様書 10章 検証チェックリスト）。"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook

from .engine import Matrix
from .importer import read_data_sheet
from .master import Master


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str = ""

    def __str__(self) -> str:
        mark = "✓" if self.passed else "✗"
        return f"[{mark}] {self.name}" + (f" — {self.detail}" if self.detail else "")


def validate_file(path: str, master: Master, expected_count: int | None = None) -> list[CheckResult]:
    """生成済みxlsxを読み戻して検証チェックリストを実行する。"""
    results: list[CheckResult] = []
    matrices = read_data_sheet(path)

    # マトリクス総数
    if expected_count is not None:
        results.append(
            CheckResult(
                "マトリクス総数",
                len(matrices) == expected_count,
                f"想定{expected_count}個 / 実際{len(matrices)}個",
            )
        )

    # 配列構造: マトリクス間が2行空いているか
    results.append(_check_gaps(path))

    # 全セルがMASTERに存在する料金ランクか
    bad = [
        (str(m.base_rank), lead, stock, str(cell))
        for m in matrices
        for lead, row in zip(m.lead_times, m.cells)
        for stock, cell in zip(m.stock_axis, row)
        if not master.contains(cell)
    ]
    results.append(
        CheckResult(
            "全セルがMASTERの料金ランク",
            not bad,
            "" if not bad else f"不正セル{len(bad)}個 例: {bad[0]}",
        )
    )

    # 料金ランク順: 同一軸グループ内でベース料金ランクがMASTER順（昇順）か
    groups: dict[tuple, list[Matrix]] = {}
    for m in matrices:
        groups.setdefault((tuple(m.lead_times), tuple(m.stock_axis)), []).append(m)
    ordered = all(
        all(a.base_rank.amount < b.base_rank.amount for a, b in zip(ms, ms[1:]))
        for ms in groups.values()
    )
    results.append(CheckResult("料金ランク順（グループ内でMASTER順）", ordered))

    # 下限打ち止め: 最低ベースのマトリクスの最小セルが最低料金ランクか
    lowest = min(matrices, key=lambda m: m.base_rank.amount)
    low_cells = [c for row in lowest.cells for c in row]
    results.append(
        CheckResult(
            "下限打ち止め",
            min(low_cells).amount >= master.min_rank.amount,
            f"最低ベース{lowest.base_rank}の最小セル: {min(low_cells)}",
        )
    )

    # 上限打ち止め: 最高ベースのマトリクスの最大セルが最高料金ランク以下か
    highest = max(matrices, key=lambda m: m.base_rank.amount)
    high_cells = [c for row in highest.cells for c in row]
    results.append(
        CheckResult(
            "上限打ち止め",
            max(high_cells).amount <= master.max_rank.amount,
            f"最高ベース{highest.base_rank}の最大セル: {max(high_cells)}",
        )
    )

    # 部屋タイプ名が全マトリクスで一致しているか
    names = {m.room_type for m in matrices}
    results.append(
        CheckResult("部屋タイプ名の一致", len(names) == 1, f"検出: {sorted(names)}")
    )

    return results


def _check_gaps(path: str) -> CheckResult:
    """DATAシートで各マトリクス間がちょうど2行空いているかを確認する。"""
    ws = load_workbook(path, data_only=True)["DATA"]
    starts = [
        row[0].row
        for row in ws.iter_rows(min_col=1, max_col=1)
        if row[0].value == "部屋タイプ"
    ]
    last_data_rows = []
    for row in ws.iter_rows(min_col=2, max_col=2):
        if row[0].value is not None and str(row[0].value) != "":
            last_data_rows.append(row[0].row)

    bad_gaps = []
    for prev_start, next_start in zip(starts, starts[1:]):
        prev_end = max(r for r in last_data_rows if r < next_start)
        gap = next_start - prev_end - 1
        if gap != 2:
            bad_gaps.append((prev_start, next_start, gap))
    return CheckResult(
        "マトリクス間が2行空き",
        not bad_gaps,
        "" if not bad_gaps else f"不正な間隔{len(bad_gaps)}箇所 例: {bad_gaps[0]}",
    )
